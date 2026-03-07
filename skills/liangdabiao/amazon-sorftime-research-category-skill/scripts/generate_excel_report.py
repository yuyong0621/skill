#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
品类选品 Excel 报告生成脚本
功能: 根据收集的数据生成包含12个sheets的Excel报告
"""

import json
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

# 样式定义
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14)
DATA_FONT = Font(size=10)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
LEFT_ALIGN = Alignment(horizontal="left", vertical="center")
BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

class CategoryReportGenerator:
    def __init__(self, category_data, output_path):
        """
        初始化报告生成器

        Args:
            category_data: dict, 包含所有收集的数据
            output_path: str, 输出文件路径
        """
        self.data = category_data
        self.output_path = output_path
        self.wb = openpyxl.Workbook()

    def generate(self):
        """生成完整报告"""
        print("开始生成 Excel 报告...")

        # 生成12个sheets
        self._create_overview_sheet()      # Sheet 1: 市场概览
        self._create_top100_sheet()        # Sheet 2: Top100产品
        self._create_brand_sheet()         # Sheet 3: 品牌分析
        self._create_price_dist_sheet()    # Sheet 4: 价格分布
        self._create_rating_dist_sheet()   # Sheet 5: 评分分布
        self._create_seller_sheet()        # Sheet 6: 卖家来源
        self._create_new_product_sheet()   # Sheet 7: 新产品分析
        self._create_trend_sales_sheet()   # Sheet 8: 销量趋势
        self._create_trend_price_sheet()   # Sheet 9: 价格趋势
        self._create_trend_score_sheet()   # Sheet 10: 评分趋势
        self._create_trend_brand_sheet()   # Sheet 11: 品牌数趋势
        self._create_supply_chain_sheet()  # Sheet 12: 供应链

        # 保存文件
        self.wb.save(self.output_path)
        print(f"Excel 报告已生成: {self.output_path}")

    def _create_overview_sheet(self):
        """Sheet 1: 市场概览 - 五维评分模型 + KPI指标"""
        ws = self.wb.active
        ws.title = "市场概览"

        row = 1

        # 报告标题
        ws.cell(row, 1, f"{self.data.get('category_name', '')} 品类市场调研报告")
        ws.cell(row, 1).font = TITLE_FONT
        row += 2

        # 五维评分模型
        ws.cell(row, 1, "五维评分模型")
        ws.cell(row, 1).font = HEADER_FONT
        ws.cell(row, 1).fill = HEADER_FILL
        row += 1

        scores = self.data.get('five_dimension_score', {})
        score_data = [
            ["评分维度", "得分", "满分", "占比"],
            ["市场规模", scores.get("market_size", 0), 20, f"{scores.get('market_size', 0)/20*100:.0f}%"],
            ["增长潜力", scores.get("growth_potential", 0), 25, f"{scores.get('growth_potential', 0)/25*100:.0f}%"],
            ["竞争烈度", scores.get("competition", 0), 20, f"{scores.get('competition', 0)/20*100:.0f}%"],
            ["进入壁垒", scores.get("entry_barrier", 0), 20, f"{scores.get('entry_barrier', 0)/20*100:.0f}%"],
            ["利润空间", scores.get("profit_margin", 0), 15, f"{scores.get('profit_margin', 0)/15*100:.0f}%"],
            ["", "", "", ""],
            ["总分", scores.get("total", 0), 100, "100%"],
        ]

        for r in score_data:
            for c, val in enumerate(r, 1):
                cell = ws.cell(row, c, val)
                cell.font = DATA_FONT
                cell.border = BORDER
                if c == 1:
                    cell.alignment = LEFT_ALIGN
                else:
                    cell.alignment = CENTER_ALIGN
            row += 1

        row += 2

        # 评级和建议
        total_score = scores.get("total", 0)
        rating = self._get_rating(total_score)
        ws.cell(row, 1, f"评级: {rating}")
        ws.cell(row, 1).font = Font(bold=True, size=12, color="FF0000" if total_score < 40 else "008000")
        row += 1
        ws.cell(row, 1, f"建议: {self._get_recommendation(total_score)}")
        row += 3

        # KPI指标
        ws.cell(row, 1, "关键指标 (KPI)")
        ws.cell(row, 1).font = HEADER_FONT
        ws.cell(row, 1).fill = HEADER_FILL
        row += 1

        kpi = self.data.get('kpi', {})
        kpi_data = [
            ["指标", "数值"],
            ["产品总数", kpi.get("total_products", 0)],
            ["平均价格", f"${kpi.get("avg_price", 0):.2f}"],
            ["平均月销量", f"{kpi.get("avg_sales", 0):.0f}"],
            ["平均评分", f"{kpi.get("avg_rating", 0):.2f}"],
            ["总月销额", f"${kpi.get("total_sales", 0):,.0f}"],
            ["CR3集中度", f"{kpi.get("cr3", 0):.2f}%"],
            ["HHI指数", f"{kpi.get("hhi", 0):.2f}"],
        ]

        for r in kpi_data:
            for c, val in enumerate(r, 1):
                cell = ws.cell(row, c, val)
                cell.font = DATA_FONT
                cell.border = BORDER
            row += 1

        self._adjust_column_width(ws)

    def _create_top100_sheet(self):
        """Sheet 2: Top100产品"""
        ws = self.wb.create_sheet(title="Top100产品")

        # 表头
        headers = ["排名", "ASIN", "产品标题", "品牌", "价格", "评分", "评论数",
                   "月销量", "月销额($)", "市场份额", "卖家", "卖家来源"]

        row = 1
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row, c, h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER_ALIGN
            cell.border = BORDER

        # 数据行
        products = self.data.get('top100_products', [])
        for idx, p in enumerate(products, 2):
            ws.cell(idx, 1, idx - 1)  # 排名
            ws.cell(idx, 2, p.get("asin", ""))
            ws.cell(idx, 3, p.get("title", ""))
            ws.cell(idx, 4, p.get("brand", ""))
            ws.cell(idx, 5, p.get("price", 0))
            ws.cell(idx, 6, p.get("rating", 0))
            ws.cell(idx, 7, p.get("review_count", 0))
            ws.cell(idx, 8, p.get("monthly_sales", 0))
            ws.cell(idx, 9, p.get("monthly_revenue", 0))
            ws.cell(idx, 10, f"{p.get("market_share", 0):.2f}%")
            ws.cell(idx, 11, p.get("seller", ""))
            ws.cell(idx, 12, p.get("seller_source", ""))

            # 应用样式
            for c in range(1, 13):
                cell = ws.cell(idx, c)
                cell.font = DATA_FONT
                cell.border = BORDER
                if c in [1]:  # 排名列居中
                    cell.alignment = CENTER_ALIGN
                elif c in [3]:  # 标题左对齐
                    cell.alignment = LEFT_ALIGN

        self._adjust_column_width(ws)

    def _create_brand_sheet(self):
        """Sheet 3: 品牌分析"""
        ws = self.wb.create_sheet(title="品牌分析")

        # 表头
        headers = ["排名", "品牌", "产品数", "月销量", "月销额($)", "市场份额", "平均评分"]

        row = 1
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row, c, h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER_ALIGN
            cell.border = BORDER

        # 数据行
        brands = self.data.get('brand_analysis', [])
        for idx, b in enumerate(brands, 2):
            ws.cell(idx, 1, idx - 1)
            ws.cell(idx, 2, b.get("brand", ""))
            ws.cell(idx, 3, b.get("product_count", 0))
            ws.cell(idx, 4, b.get("monthly_sales", 0))
            ws.cell(idx, 5, b.get("monthly_revenue", 0))
            ws.cell(idx, 6, f"{b.get("market_share", 0):.2f}%")
            ws.cell(idx, 7, b.get("avg_rating", 0))

            for c in range(1, 8):
                cell = ws.cell(idx, c)
                cell.font = DATA_FONT
                cell.border = BORDER
                cell.alignment = CENTER_ALIGN

        self._adjust_column_width(ws)

    def _create_price_dist_sheet(self):
        """Sheet 4: 价格分布"""
        ws = self.wb.create_sheet(title="价格分布")

        headers = ["价格区间", "产品数", "占比", "销量", "销额", "平均评分"]

        row = 1
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row, c, h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER_ALIGN
            cell.border = BORDER

        price_dist = self.data.get('price_distribution', [])
        for idx, d in enumerate(price_dist, 2):
            ws.cell(idx, 1, d.get("range", ""))
            ws.cell(idx, 2, d.get("count", 0))
            ws.cell(idx, 3, f"{d.get("percentage", 0):.1f}%")
            ws.cell(idx, 4, d.get("sales", 0))
            ws.cell(idx, 5, d.get("revenue", 0))
            ws.cell(idx, 6, d.get("avg_rating", 0))

            for c in range(1, 7):
                cell = ws.cell(idx, c)
                cell.font = DATA_FONT
                cell.border = BORDER
                cell.alignment = CENTER_ALIGN

        self._adjust_column_width(ws)

    def _create_rating_dist_sheet(self):
        """Sheet 5: 评分分布"""
        ws = self.wb.create_sheet(title="评分分布")

        headers = ["评分区间", "产品数", "占比", "销量占比"]

        row = 1
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row, c, h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER_ALIGN
            cell.border = BORDER

        rating_dist = self.data.get('rating_distribution', [])
        for idx, d in enumerate(rating_dist, 2):
            ws.cell(idx, 1, d.get("range", ""))
            ws.cell(idx, 2, d.get("count", 0))
            ws.cell(idx, 3, f"{d.get("percentage", 0):.1f}%")
            ws.cell(idx, 4, f"{d.get("sales_percentage", 0):.1f}%")

            for c in range(1, 5):
                cell = ws.cell(idx, c)
                cell.font = DATA_FONT
                cell.border = BORDER
                cell.alignment = CENTER_ALIGN

        self._adjust_column_width(ws)

    def _create_seller_sheet(self):
        """Sheet 6: 卖家来源"""
        ws = self.wb.create_sheet(title="卖家来源")

        headers = ["来源地", "卖家数", "产品数", "占比", "销额"]

        row = 1
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row, c, h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER_ALIGN
            cell.border = BORDER

        seller_dist = self.data.get('seller_distribution', [])
        for idx, d in enumerate(seller_dist, 2):
            ws.cell(idx, 1, d.get("source", ""))
            ws.cell(idx, 2, d.get("seller_count", 0))
            ws.cell(idx, 3, d.get("product_count", 0))
            ws.cell(idx, 4, f"{d.get("percentage", 0):.1f}%")
            ws.cell(idx, 5, d.get("revenue", 0))

            for c in range(1, 6):
                cell = ws.cell(idx, c)
                cell.font = DATA_FONT
                cell.border = BORDER
                cell.alignment = CENTER_ALIGN

        self._adjust_column_width(ws)

    def _create_new_product_sheet(self):
        """Sheet 7: 新产品分析"""
        ws = self.wb.create_sheet(title="新产品分析")

        headers = ["ASIN", "产品标题", "品牌", "价格", "评分", "评论数", "月销量", "上架天数"]

        row = 1
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row, c, h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER_ALIGN
            cell.border = BORDER

        new_products = self.data.get('new_products', [])
        for idx, p in enumerate(new_products, 2):
            ws.cell(idx, 1, p.get("asin", ""))
            ws.cell(idx, 2, p.get("title", ""))
            ws.cell(idx, 3, p.get("brand", ""))
            ws.cell(idx, 4, p.get("price", 0))
            ws.cell(idx, 5, p.get("rating", 0))
            ws.cell(idx, 6, p.get("review_count", 0))
            ws.cell(idx, 7, p.get("monthly_sales", 0))
            ws.cell(idx, 8, p.get("days_online", 0))

            for c in range(1, 9):
                cell = ws.cell(idx, c)
                cell.font = DATA_FONT
                cell.border = BORDER
                if c in [1, 8]:  # ASIN和上架天数居中
                    cell.alignment = CENTER_ALIGN
                elif c in [2]:  # 标题左对齐
                    cell.alignment = LEFT_ALIGN

        self._adjust_column_width(ws)

    def _create_trend_sales_sheet(self):
        """Sheet 8: 销量趋势"""
        ws = self.wb.create_sheet(title="趋势-销量")

        headers = ["日期", "月销量", "环比", "同比"]

        row = 1
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row, c, h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER_ALIGN
            cell.border = BORDER

        trend = self.data.get('sales_trend', [])
        for idx, t in enumerate(trend, 2):
            ws.cell(idx, 1, t.get("date", ""))
            ws.cell(idx, 2, t.get("sales", 0))
            ws.cell(idx, 3, f"{t.get("mom", 0):.1f}%")
            ws.cell(idx, 4, f"{t.get("yoy", 0):.1f}%")

            for c in range(1, 5):
                cell = ws.cell(idx, c)
                cell.font = DATA_FONT
                cell.border = BORDER
                cell.alignment = CENTER_ALIGN

        self._adjust_column_width(ws)

    def _create_trend_price_sheet(self):
        """Sheet 9: 价格趋势"""
        ws = self.wb.create_sheet(title="趋势-价格")

        headers = ["日期", "平均价格", "环比"]

        row = 1
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row, c, h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER_ALIGN
            cell.border = BORDER

        trend = self.data.get('price_trend', [])
        for idx, t in enumerate(trend, 2):
            ws.cell(idx, 1, t.get("date", ""))
            ws.cell(idx, 2, f"${t.get("avg_price", 0):.2f}")
            ws.cell(idx, 3, f"{t.get("mom", 0):.1f}%")

            for c in range(1, 4):
                cell = ws.cell(idx, c)
                cell.font = DATA_FONT
                cell.border = BORDER
                cell.alignment = CENTER_ALIGN

        self._adjust_column_width(ws)

    def _create_trend_score_sheet(self):
        """Sheet 10: 评分趋势"""
        ws = self.wb.create_sheet(title="趋势-评分")

        headers = ["日期", "平均评分", "环比"]

        row = 1
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row, c, h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER_ALIGN
            cell.border = BORDER

        trend = self.data.get('rating_trend', [])
        for idx, t in enumerate(trend, 2):
            ws.cell(idx, 1, t.get("date", ""))
            ws.cell(idx, 2, f"{t.get("avg_score", 0):.2f}")
            ws.cell(idx, 3, f"{t.get("mom", 0):.1f}%")

            for c in range(1, 4):
                cell = ws.cell(idx, c)
                cell.font = DATA_FONT
                cell.border = BORDER
                cell.alignment = CENTER_ALIGN

        self._adjust_column_width(ws)

    def _create_trend_brand_sheet(self):
        """Sheet 11: 品牌数趋势"""
        ws = self.wb.create_sheet(title="趋势-品牌数")

        headers = ["日期", "品牌数", "新品牌数", "退出品牌数"]

        row = 1
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row, c, h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER_ALIGN
            cell.border = BORDER

        trend = self.data.get('brand_count_trend', [])
        for idx, t in enumerate(trend, 2):
            ws.cell(idx, 1, t.get("date", ""))
            ws.cell(idx, 2, t.get("brand_count", 0))
            ws.cell(idx, 3, t.get("new_brands", 0))
            ws.cell(idx, 4, t.get("exited_brands", 0))

            for c in range(1, 5):
                cell = ws.cell(idx, c)
                cell.font = DATA_FONT
                cell.border = BORDER
                cell.alignment = CENTER_ALIGN

        self._adjust_column_width(ws)

    def _create_supply_chain_sheet(self):
        """Sheet 12: 供应链-1688"""
        ws = self.wb.create_sheet(title="供应链-1688")

        headers = ["产品名称", "1688价格", "Amazon均价", "预估毛利率", "1688链接"]

        row = 1
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row, c, h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER_ALIGN
            cell.border = BORDER

        supply = self.data.get('supply_chain_1688', [])
        for idx, s in enumerate(supply, 2):
            ws.cell(idx, 1, s.get("product_name", ""))
            ws.cell(idx, 2, f"¥{s.get('price_1688', 0):.2f}")
            ws.cell(idx, 3, f"${s.get('price_amazon', 0):.2f}")
            ws.cell(idx, 4, f"{s.get('margin', 0):.1f}%")
            ws.cell(idx, 5, s.get("link", ""))

            for c in range(1, 6):
                cell = ws.cell(idx, c)
                cell.font = DATA_FONT
                cell.border = BORDER
                if c in [5]:  # 链接列左对齐
                    cell.alignment = LEFT_ALIGN
                else:
                    cell.alignment = CENTER_ALIGN

        self._adjust_column_width(ws)

    def _get_rating(self, score):
        """根据分数获取评级"""
        if score >= 80:
            return "优秀"
        elif score >= 60:
            return "良好"
        elif score >= 40:
            return "一般"
        else:
            return "较差"

    def _get_recommendation(self, score):
        """根据分数获取建议"""
        if score >= 80:
            return "强烈推荐进入"
        elif score >= 60:
            return "可以考虑进入"
        elif score >= 40:
            return "谨慎进入"
        else:
            return "不建议进入"

    def _adjust_column_width(self, ws):
        """自动调整列宽"""
        column_widths = {
            'A': 8, 'B': 15, 'C': 40, 'D': 15, 'E': 10, 'F': 10,
            'G': 12, 'H': 12, 'I': 15, 'J': 12, 'K': 15, 'L': 12
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width


def main():
    """示例用法"""
    # 示例数据结构
    category_data = {
        "category_name": "Sofas",
        "five_dimension_score": {
            "market_size": 18,
            "growth_potential": 22,
            "competition": 15,
            "entry_barrier": 16,
            "profit_margin": 11,
            "total": 82
        },
        "kpi": {
            "total_products": 100,
            "avg_price": 125.50,
            "avg_sales": 1245,
            "avg_rating": 4.2,
            "total_sales": 12500000,
            "cr3": 12.39,
            "hhi": 167.71
        },
        "top100_products": [],
        "brand_analysis": [],
        "price_distribution": [],
        "rating_distribution": [],
        "seller_distribution": [],
        "new_products": [],
        "sales_trend": [],
        "price_trend": [],
        "rating_trend": [],
        "brand_count_trend": [],
        "supply_chain_1688": []
    }

    generator = CategoryReportGenerator(
        category_data=category_data,
        output_path="category-reports/2026/03/category_sofas_US_20260303.xlsx"
    )
    generator.generate()


if __name__ == "__main__":
    main()
