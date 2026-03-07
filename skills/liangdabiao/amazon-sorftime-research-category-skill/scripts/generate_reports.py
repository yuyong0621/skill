#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
品类选品分析统一报告生成器 v3.0
支持: Markdown, Excel, HTML, CSV 四种格式
所有报告保存到统一的输出文件夹

主要改进 (v3.0):
1. 完整的模板变量替换（包括评级、分析、策略等）
2. 自动生成分析文本和风险提示
3. 修复 f-string 花括号转义问题
4. 支持 Jinja2 风格模板语法
5. 新增品牌分析表格
"""

import os
import sys
import json
import csv
import shutil
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional


class CategoryReportGenerator:
    """品类选品分析报告生成器"""

    def __init__(self, data: Dict, output_dir: Optional[str] = None):
        """
        初始化报告生成器

        Args:
            data: 包含 statistics, products, scores 的字典
            output_dir: 输出目录，默认为 category-reports/{品类名}_{日期}/
        """
        self.data = data
        self.statistics = data.get('statistics', {})
        self.products = data.get('products', [])
        self.scores = data.get('scores', {})

        # 确定输出目录
        if output_dir:
            self.output_dir = Path(output_dir)
        else:
            # 使用品类名和日期创建文件夹
            category_name = self._get_category_name()
            date_str = datetime.now().strftime('%Y%m%d')
            safe_name = self._sanitize_filename(category_name)
            self.output_dir = Path('category-reports') / f"{safe_name}_{date_str}"

        # 创建输出目录
        self.output_dir.mkdir(parents=True, exist_ok=True)

        # 创建数据子目录
        self.data_dir = self.output_dir / 'data'
        self.data_dir.mkdir(exist_ok=True)

        self.generated_files = []

    def _get_category_name(self) -> str:
        """从数据中获取品类名称"""
        # 尝试从各个位置获取品类名
        if 'category_name' in self.data:
            return self.data['category_name']
        if '品类名称' in self.data:
            return self.data['品类名称']
        return 'Unknown_Category'

    def _sanitize_filename(self, name: str) -> str:
        """清理文件名，移除非法字符"""
        # 移除或替换非法字符
        illegal_chars = '<>:"/\\|?*'
        for char in illegal_chars:
            name = name.replace(char, '_')
        # 移除空格
        name = name.replace(' ', '_')
        return name[:100]  # 限制长度

    def generate_all(self) -> Dict[str, str]:
        """
        生成所有格式的报告

        Returns:
            生成的文件路径字典 {格式: 文件路径}
        """
        results = {}

        # 1. Markdown 报告
        try:
            md_path = self.generate_markdown()
            results['markdown'] = str(md_path)
        except Exception as e:
            print(f"Markdown 生成失败: {e}")

        # 2. CSV 数据文件
        try:
            csv_path = self.generate_csv()
            results['csv'] = str(csv_path)
        except Exception as e:
            print(f"CSV 生成失败: {e}")

        # 3. Excel 报告 (需要 openpyxl)
        try:
            excel_path = self.generate_excel()
            results['excel'] = str(excel_path)
        except ImportError:
            print("Excel 生成跳过 (需要 openpyxl)")
        except Exception as e:
            print(f"Excel 生成失败: {e}")

        # 4. HTML 仪表板
        try:
            html_path = self.generate_html()
            results['html'] = str(html_path)
        except Exception as e:
            print(f"HTML 生成失败: {e}")

        # 5. 保存原始 JSON 数据
        try:
            json_path = self.generate_json()
            results['json'] = str(json_path)
        except Exception as e:
            print(f"JSON 生成失败: {e}")

        return results

    def generate_markdown(self) -> Path:
        """生成 Markdown 报告"""
        filename = self.output_dir / "category_analysis_report.md"

        # 读取模板
        template_path = Path(__file__).parent.parent / 'assets' / 'report_template.md'
        if template_path.exists():
            with open(template_path, 'r', encoding='utf-8') as f:
                content = f.read()
        else:
            content = self._get_default_markdown_template()

        # 替换变量
        content = self._replace_variables(content)

        # 写入文件
        with open(filename, 'w', encoding='utf-8') as f:
            f.write(content)

        self.generated_files.append(filename)
        print(f"✓ Markdown 报告: {filename}")
        return filename

    def generate_csv(self) -> Path:
        """生成 CSV 数据文件"""
        # 1. 统计数据 CSV
        stats_file = self.data_dir / "statistics.csv"
        with open(stats_file, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            writer.writerow(['指标', '数值'])
            for key, value in self.statistics.items():
                writer.writerow([key, value])

        # 2. 产品列表 CSV
        products_file = self.data_dir / "products.csv"
        with open(products_file, 'w', newline='', encoding='utf-8-sig') as f:
            fieldnames = ['ASIN', '标题', '品牌', '价格', '月销量', '评分', '排名']
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            for i, p in enumerate(self.products, 1):
                writer.writerow({
                    'ASIN': p.get('ASIN', ''),
                    '标题': p.get('标题', '')[:100],
                    '品牌': p.get('品牌', ''),
                    '价格': p.get('价格', 0),
                    '月销量': p.get('月销量', 0),
                    '评分': p.get('评分', 0),
                    '排名': i
                })

        # 3. 评分数据 CSV
        scores_file = self.data_dir / "scores.csv"
        with open(scores_file, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            writer.writerow(['维度', '得分', '满分', '占比%'])
            max_scores = {'市场规模': 20, '增长潜力': 25, '竞争烈度': 20, '进入壁垒': 20, '利润空间': 15}
            total_max = sum(max_scores.values())
            for key, value in self.scores.items():
                if key in max_scores:
                    max_score = max_scores[key]
                    pct = (value / max_score * 100) if max_score > 0 else 0
                    writer.writerow([key, value, max_score, f"{pct:.1f}"])
                elif key not in ['总分', '评级']:
                    writer.writerow([key, value, '', ''])

        self.generated_files.extend([stats_file, products_file, scores_file])
        print(f"✓ CSV 数据文件: {self.data_dir}")
        return stats_file  # 返回主文件

    def generate_excel(self) -> Path:
        """生成 Excel 报告"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font as OpenpyxlFont, Alignment, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError("需要安装 openpyxl: pip install openpyxl")

        filename = self.output_dir / "category_analysis_report.xlsx"
        wb = Workbook()

        # 删除默认工作表
        wb.remove(wb.active)

        # 1. 概览工作表
        self._create_overview_sheet(wb, OpenpyxlFont)

        # 2. 产品列表工作表
        self._create_products_sheet(wb, OpenpyxlFont, PatternFill)

        # 3. 评分工作表
        self._create_scores_sheet(wb, OpenpyxlFont)

        wb.save(filename)
        self.generated_files.append(filename)
        print(f"✓ Excel 报告: {filename}")
        return filename

    def generate_html(self) -> Path:
        """生成 HTML 仪表板"""
        filename = self.output_dir / "dashboard.html"

        # 读取模板
        template_path = Path(__file__).parent.parent / 'assets' / 'dashboard_template.html'
        if template_path.exists():
            with open(template_path, 'r', encoding='utf-8') as f:
                content = f.read()
        else:
            content = self._get_default_html_template()

        # 替换变量
        content = self._replace_html_variables(content)

        with open(filename, 'w', encoding='utf-8') as f:
            f.write(content)

        self.generated_files.append(filename)
        print(f"✓ HTML 仪表板: {filename}")
        return filename

    def generate_json(self) -> Path:
        """保存原始 JSON 数据"""
        filename = self.data_dir / "raw_data.json"
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(self.data, f, ensure_ascii=False, indent=2)

        self.generated_files.append(filename)
        print(f"✓ JSON 数据: {filename}")
        return filename

    def _replace_variables(self, content: str) -> str:
        """替换 Markdown 模板变量"""
        # 基础信息
        content = content.replace('{{DATE}}', datetime.now().strftime('%Y-%m-%d'))
        content = content.replace('{{TIME}}', datetime.now().strftime('%H:%M:%S'))
        content = content.replace('{{CATEGORY_NAME}}', self._get_category_name())

        # 评级和建议
        total_score = self.scores.get('总分', 0)
        if total_score >= 80:
            rating = "优秀"
            recommendation = "强烈推荐进入"
        elif total_score >= 70:
            rating = "良好"
            recommendation = "可以考虑进入"
        elif total_score >= 50:
            rating = "一般"
            recommendation = "谨慎进入"
        else:
            rating = "较差"
            recommendation = "不建议进入"

        content = content.replace('{{SCORE_评级}}', rating)
        content = content.replace('{{SCORE_建议}}', recommendation)

        # 统计数据表格
        if '{% for key, value in statistics.items() %}' in content:
            # 替换 Jinja2 风格的循环为实际表格
            stats_table = ""
            for key, value in self.statistics.items():
                stats_table += f"| {key} | {value} |\n"
            content = content.replace('{% for key, value in statistics.items() %}', '')
            content = content.replace('| {{ key }} | {{ value }} |', stats_table)
            content = content.replace('{% endfor %}', '')

        # 统计数据变量
        for key, value in self.statistics.items():
            content = content.replace(f'{{{{STAT_{key}}}}}', str(value))

        # 评分 (包括占比计算)
        max_scores = {'市场规模': 20, '增长潜力': 25, '竞争烈度': 20, '进入壁垒': 20, '利润空间': 15}
        for key, value in self.scores.items():
            placeholder = '{{' + f'SCORE_{key}' + '}}'
            content = content.replace(placeholder, str(value))
            # 添加占比
            if key in max_scores:
                pct = (value / max_scores[key] * 100) if max_scores[key] > 0 else 0
                placeholder_pct = '{{' + f'SCORE_{key}_占比' + '}}'
                content = content.replace(placeholder_pct, f'{pct:.1f}%')

        # 生成分析文本
        analysis = self._generate_analysis_text()
        for key, value in analysis.items():
            content = content.replace(f'{{{{ANALYSIS_{key}}}}}', value)

        # 生成策略文本
        strategies = self._generate_strategy_text()
        for key, value in strategies.items():
            content = content.replace(f'{{{{STRATEGY_{key}}}}}', value)

        # 风险提示
        risks = self._generate_risk_text()
        content = content.replace('{{RISK_提示}}', risks)

        # 产品列表
        if '{{PRODUCTS_TABLE}}' in content and self.products:
            products_table = "| 排名 | ASIN | 品牌 | 价格 | 月销量 | 评分 | 标题 |\n"
            products_table += "|------|------|------|------|--------|------|------|\n"
            for i, p in enumerate(self.products[:20], 1):
                title = p.get('标题', '')[:50]
                # 安全获取数值
                price = self._safe_float(p.get('价格', 0))
                sales = self._safe_int(p.get('月销量', 0))
                rating = self._safe_float(p.get('评分', 0))

                products_table += f"| {i} | {p.get('ASIN', '')} | {p.get('品牌', '')} | "
                products_table += f"${price:.2f} | {sales:,} | "
                products_table += f"{rating:.1f} | {title}... |\n"
            content = content.replace('{{PRODUCTS_TABLE}}', products_table)

        # 品牌表格
        if '{{BRANDS_TABLE}}' in content:
            brands_table = self._generate_brands_table()
            content = content.replace('{{BRANDS_TABLE}}', brands_table)

        return content

    def _generate_analysis_text(self) -> dict:
        """生成分析文本"""
        total_revenue = self._safe_float(self.statistics.get('总销额', 0))
        avg_price = self._safe_float(self.statistics.get('平均价格', 0))
        top3_share = self._safe_float(self.statistics.get('Top3品牌占比', 0))

        analysis = {}

        # 市场规模分析
        if total_revenue > 50_000_000:
            analysis['市场规模'] = "类目月销额超过 5000 万美元，属于大规模市场，具有巨大的销售潜力。"
        elif total_revenue > 10_000_000:
            analysis['市场规模'] = "类目月销额在 1000 万至 5000 万美元之间，市场规模可观。"
        else:
            analysis['市场规模'] = "类目月销额低于 1000 万美元，市场规模相对较小。"

        # 增长潜力分析
        growth_score = self.scores.get('增长潜力', 0)
        if growth_score >= 20:
            analysis['增长潜力'] = "低评论产品占比较高，说明新品有较大的市场机会，增长潜力良好。"
        else:
            analysis['增长潜力'] = "低评论产品占比较低，市场竞争激烈，新品增长机会有限。"

        # 竞争烈度分析
        if top3_share < 30:
            analysis['竞争烈度'] = "Top3 品牌占比较低，市场竞争相对分散，新进入者有较多机会。"
        elif top3_share < 50:
            analysis['竞争烈度'] = "Top3 品牌占比适中，市场存在一定竞争，但仍有机会。"
        else:
            analysis['竞争烈度'] = f"Top3 品牌占比达到 {top3_share:.1f}%，市场高度集中，竞争非常激烈。"

        # 进入壁垒分析
        barrier_score = self.scores.get('进入壁垒', 0)
        if barrier_score >= 16:
            analysis['进入壁垒'] = "Amazon 自营占比较低且新品机会多，进入壁垒较低。"
        elif barrier_score >= 10:
            analysis['进入壁垒'] = "Amazon 自营占比适中或新品机会一般，进入壁垒中等。"
        else:
            analysis['进入壁垒'] = "Amazon 自营占比较高且新品机会少，进入壁垒较高。"

        # 利润空间分析
        profit_score = self.scores.get('利润空间', 0)
        if profit_score >= 10:
            analysis['利润空间'] = f"平均价格 ${avg_price:.0f}，利润空间充足。"
        elif profit_score >= 6:
            analysis['利润空间'] = f"平均价格 ${avg_price:.0f}，利润空间一般。"
        else:
            analysis['利润空间'] = f"平均价格 ${avg_price:.0f}，利润空间有限，需注重成本控制。"

        return analysis

    def _generate_strategy_text(self) -> dict:
        """生成策略建议"""
        strategies = {}
        avg_price = self._safe_float(self.statistics.get('平均价格', 0))

        # 进入策略
        total_score = self.scores.get('总分', 0)
        if total_score >= 70:
            strategies['进入'] = "该品类综合评分良好，可以考虑进入。建议关注差异化产品和细分市场。"
        elif total_score >= 50:
            strategies['进入'] = "该品类综合评分一般，建议谨慎进入。需做好充分的市场调研和竞争分析。"
        else:
            strategies['进入'] = "该品类综合评分较低，不建议进入。建议考虑其他品类。"

        # 定位建议
        strategies['定位'] = f"建议定价在 ${avg_price*0.8:.0f} - ${avg_price*1.2:.0f} 范围内，注重产品质量和差异化功能。"

        # 定价建议
        if avg_price > 300:
            strategies['定价'] = "高端市场定价策略，强调品质和品牌价值。"
        elif avg_price > 100:
            strategies['定价'] = "中端市场定价策略，平衡性价比。"
        else:
            strategies['定价'] = "经济型定价策略，注重成本控制和销量规模。"

        return strategies

    def _generate_risk_text(self) -> str:
        """生成风险提示"""
        risks = []
        top3_share = self._safe_float(self.statistics.get('Top3品牌占比', 0))

        if top3_share > 50:
            risks.append(f"- 市场高度集中，Top3 品牌占比 {top3_share:.1f}%，难以与头部品牌竞争")

        if self.scores.get('进入壁垒', 0) < 10:
            risks.append("- Amazon 自营占比较高，需注意价格竞争")

        if self.scores.get('增长潜力', 0) < 15:
            risks.append("- 新品机会有限，需投入更多营销资源")

        if not risks:
            risks.append("- 需关注产品质量以获取好评")
            risks.append("- 建议定期跟踪市场变化")

        return "\n".join(risks)

    def _generate_brands_table(self) -> str:
        """生成品牌分析表格"""
        if not self.products:
            return "| 暂无数据 |\n"

        # 统计品牌数据
        brand_stats = {}
        for p in self.products[:20]:
            brand = p.get('品牌', 'Unknown')
            if brand not in brand_stats:
                brand_stats[brand] = {'count': 0, 'total_price': 0, 'total_rating': 0}
            brand_stats[brand]['count'] += 1
            brand_stats[brand]['total_price'] += self._safe_float(p.get('价格', 0))
            brand_stats[brand]['total_rating'] += self._safe_float(p.get('评分', 0))

        # 生成表格
        table = "| 品牌 | 产品数 | 平均价格 | 平均评分 |\n"
        table += "|------|--------|----------|----------|\n"

        for brand, stats in sorted(brand_stats.items(), key=lambda x: x[1]['count'], reverse=True)[:10]:
            avg_price = stats['total_price'] / stats['count'] if stats['count'] > 0 else 0
            avg_rating = stats['total_rating'] / stats['count'] if stats['count'] > 0 else 0
            table += f"| {brand} | {stats['count']} | ${avg_price:.2f} | {avg_rating:.1f} |\n"

        return table

    def _safe_float(self, value) -> float:
        """安全转换为浮点数"""
        try:
            return float(str(value).replace(',', '').replace('%', ''))
        except:
            return 0.0

    def _safe_int(self, value) -> int:
        """安全转换为整数"""
        try:
            return int(float(str(value).replace(',', '').replace('%', '')))
        except:
            return 0

    def _replace_html_variables(self, content: str) -> str:
        """替换 HTML 模板变量"""
        now = datetime.now()

        # ===== 基础信息 =====
        content = content.replace('{{CATEGORY_NAME}}', self._get_category_name())
        content = content.replace('{{SITE}}', self.data.get('site', 'US'))
        content = content.replace('{{DATA_DATE}}', now.strftime('%Y-%m-%d'))
        content = content.replace('{{GENERATED_TIME}}', now.strftime('%Y-%m-%d %H:%M:%S'))
        content = content.replace('{{DATE}}', now.strftime('%Y-%m-%d'))  # 兼容旧模板

        # ===== 五维评分 (计算百分比) =====
        max_scores = {
            '市场规模': 20,
            '增长潜力': 25,
            '竞争烈度': 20,
            '进入壁垒': 20,
            '利润空间': 15
        }

        # 英文键名映射
        score_key_map = {
            'market_size': '市场规模',
            'growth_potential': '增长潜力',
            'competition': '竞争烈度',
            'entry_barrier': '进入壁垒',
            'profit_margin': '利润空间'
        }

        # 处理中文键名评分
        for cn_key, max_score in max_scores.items():
            score = self.scores.get(cn_key, 0)
            percent = (score / max_score * 100) if max_score > 0 else 0

            # 变量名转换: 市场规模 -> MARKET_SIZE
            var_key = cn_key.upper().replace('潜力', '_POTENTIAL').replace('烈度', '').replace('壁垒', '_BARRIER').replace('空间', '_MARGIN')
            if cn_key == '竞争烈度':
                var_key = 'COMPETITION'
            elif cn_key == '进入壁垒':
                var_key = 'ENTRY_BARRIER'
            elif cn_key == '利润空间':
                var_key = 'PROFIT_MARGIN'

            content = content.replace(f'{{{{{var_key}_SCORE}}}}', str(score))
            content = content.replace(f'{{{{{var_key}_PERCENT}}}}', f'{percent:.0f}')

        # 处理英文键名评分 (兼容 data_adapter 输出)
        for en_key, cn_key in score_key_map.items():
            score = self.scores.get(en_key, self.scores.get(cn_key, 0))
            max_score = max_scores[cn_key]
            percent = (score / max_score * 100) if max_score > 0 else 0

            var_key = en_key.upper()
            content = content.replace(f'{{{{{var_key}_SCORE}}}}', str(score))
            content = content.replace(f'{{{{{var_key}_PERCENT}}}}', f'{percent:.0f}')

        # 总分和评级
        total_score = self.scores.get('总分', self.scores.get('total', 0))
        rating = self.scores.get('评级', self.scores.get('rating', ''))
        content = content.replace('{{TOTAL_SCORE}}', str(total_score))
        content = content.replace('{{RATING}}', str(rating))

        # 推荐
        recommendation = self._get_recommendation(rating)
        content = content.replace('{{RECOMMENDATION}}', recommendation['text'])
        content = content.replace('{{STRATEGY}}', recommendation['strategy'])

        # ===== KPI 指标 =====
        kpi = self._calculate_kpi()
        content = content.replace('{{TOTAL_PRODUCTS}}', str(kpi['total_products']))
        content = content.replace('{{AVG_PRICE}}', f"{kpi['avg_price']:.2f}")
        content = content.replace('{{AVG_SALES}}', f"{kpi['avg_sales']:,.0f}")
        content = content.replace('{{AVG_RATING}}', f"{kpi['avg_rating']:.2f}")
        content = content.replace('{{TOTAL_SALES}}', f"{kpi['total_sales']:,.0f}")
        content = content.replace('{{CR3}}', f"{kpi['cr3']:.1f}")
        content = content.replace('{{CR3_RAW}}', f"{kpi['cr3']:.1f}")
        content = content.replace('{{HHI}}', f"{kpi['hhi']:.0f}")

        # ===== 关键发现变量 =====
        concentration_level = self._get_concentration_level(kpi['hhi'])
        content = content.replace('{{CONCENTRATION_LEVEL}}', concentration_level)

        conclusion_cr3 = self._get_cr3_conclusion(kpi['cr3'])
        content = content.replace('{{CONCLUSION_CR3}}', conclusion_cr3)

        brand_count = len(set(p.get('品牌', p.get('brand', '')) for p in self.products))
        content = content.replace('{{BRAND_COUNT}}', str(brand_count))
        content = content.replace('{{BRAND_DIVERSITY}}', '品牌多样性高' if brand_count > 20 else '品牌较集中')

        new_product_pct = kpi.get('new_product_percent', 0)
        content = content.replace('{{NEW_PRODUCT_PERCENT}}', f"{new_product_pct:.1f}")
        content = content.replace('{{NEW_PRODUCT_CONCLUSION}}', self._get_new_product_conclusion(new_product_pct))

        seller_dist = self._get_seller_distribution_summary()
        content = content.replace('{{SELLER_DISTRIBUTION}}', seller_dist)

        competition_conclusion = self._get_competition_conclusion(kpi, rating)
        content = content.replace('{{COMPETITION_CONCLUSION}}', competition_conclusion)

        # ===== 图表数据 (JavaScript JSON) =====
        chart_data = self._prepare_chart_data()

        content = content.replace('{{SALES_TREND_DATA}}', json.dumps(chart_data['sales_trend'], ensure_ascii=False))
        content = content.replace('{{PRICE_TREND_DATA}}', json.dumps(chart_data['price_trend'], ensure_ascii=False))
        content = content.replace('{{PRICE_DIST_DATA}}', json.dumps(chart_data['price_dist'], ensure_ascii=False))
        content = content.replace('{{RATING_DIST_DATA}}', json.dumps(chart_data['rating_dist'], ensure_ascii=False))
        content = content.replace('{{BRAND_SHARE_DATA}}', json.dumps(chart_data['brand_share'], ensure_ascii=False))
        content = content.replace('{{SELLER_SOURCE_DATA}}', json.dumps(chart_data['seller_source'], ensure_ascii=False))
        content = content.replace('{{BRAND_RATING_TREND_DATA}}', json.dumps(chart_data['brand_rating_trend'], ensure_ascii=False))
        content = content.replace('{{TOP50_PRODUCTS}}', json.dumps(chart_data['top50_products'], ensure_ascii=False))

        # 兼容旧模板变量
        content = content.replace('{{STATISTICS_JSON}}', json.dumps(self.statistics, ensure_ascii=False))
        content = content.replace('{{PRODUCTS_JSON}}', json.dumps(self.products[:50], ensure_ascii=False))
        content = content.replace('{{SCORES_JSON}}', json.dumps(self.scores, ensure_ascii=False))

        return content

    def _calculate_kpi(self) -> Dict:
        """计算 KPI 指标"""
        if not self.products:
            return {
                'total_products': 0,
                'avg_price': 0,
                'avg_sales': 0,
                'avg_rating': 0,
                'total_sales': 0,
                'cr3': 0,
                'hhi': 0,
                'new_product_percent': 0
            }

        total_products = len(self.products)
        avg_price = sum(self._safe_float(p.get('价格', p.get('price', 0))) for p in self.products) / total_products
        avg_sales = sum(self._safe_int(p.get('月销量', p.get('monthly_sales', 0))) for p in self.products) / total_products
        avg_rating = sum(self._safe_float(p.get('评分', p.get('rating', 0))) for p in self.products) / total_products
        total_sales = sum(self._safe_int(p.get('月销量', p.get('monthly_sales', 0))) for p in self.products)

        # 计算品牌市场份额
        brands = {}
        for p in self.products:
            brand = p.get('品牌', p.get('brand', 'Unknown'))
            revenue = self._safe_float(p.get('月销额', p.get('monthly_revenue', 0)))
            if revenue == 0:
                revenue = self._safe_float(p.get('价格', p.get('price', 0))) * self._safe_int(p.get('月销量', p.get('monthly_sales', 0)))
            brands[brand] = brands.get(brand, 0) + revenue

        total_revenue = sum(brands.values())
        if total_revenue > 0:
            brand_shares = sorted(brands.values(), reverse=True)
            cr3 = sum(brand_shares[:3]) / total_revenue * 100
        else:
            cr3 = 0

        # 计算 HHI
        if total_revenue > 0:
            hhi = sum((share / total_revenue * 100) ** 2 for share in brands.values())
        else:
            hhi = 0

        # 新品占比 (评论数 < 100)
        new_products = [p for p in self.products if self._safe_int(p.get('评论数', p.get('review_count', 0))) < 100]
        new_product_percent = len(new_products) / total_products * 100 if total_products > 0 else 0

        return {
            'total_products': total_products,
            'avg_price': avg_price,
            'avg_sales': avg_sales,
            'avg_rating': avg_rating,
            'total_sales': total_sales,
            'cr3': cr3,
            'hhi': hhi,
            'new_product_percent': new_product_percent
        }

    def _get_recommendation(self, rating: str) -> Dict[str, str]:
        """根据评级获取推荐"""
        recommendations = {
            '优秀': {
                'text': '强烈推荐进入该品类',
                'strategy': '建议尽快进入，抢占市场先机。注重产品差异化和服务质量。'
            },
            '良好': {
                'text': '可以考虑进入该品类',
                'strategy': '建议谨慎进入，寻找细分市场机会。做好竞争准备。'
            },
            '一般': {
                'text': '谨慎进入该品类',
                'strategy': '需充分评估风险，寻找差异化切入点。建议小规模试水。'
            },
            '较差': {
                'text': '不建议进入该品类',
                'strategy': '建议选择其他更有竞争力的品类，或等待市场变化。'
            }
        }
        return recommendations.get(rating, recommendations['一般'])

    def _get_concentration_level(self, hhi: float) -> str:
        """根据 HHI 获取市场集中度"""
        if hhi < 1500:
            return '低度集中'
        elif hhi < 2500:
            return '中度集中'
        else:
            return '高度集中'

    def _get_cr3_conclusion(self, cr3: float) -> str:
        """根据 CR3 获取结论"""
        if cr3 < 30:
            return '市场较为分散，新进入者机会较大'
        elif cr3 < 50:
            return '市场有一定集中度，需避开头部品牌优势领域'
        else:
            return '头部品牌优势明显，进入难度较大'

    def _get_new_product_conclusion(self, percent: float) -> str:
        """根据新品占比获取结论"""
        if percent > 40:
            return '新品表现活跃，市场对新进入者接受度高'
        elif percent > 20:
            return '新品有一定表现空间'
        else:
            return '新品表现乏力，市场存量竞争激烈'

    def _get_seller_distribution_summary(self) -> str:
        """获取卖家分布摘要"""
        sources = {}
        for p in self.products:
            source = p.get('卖家来源', p.get('seller_source', '其他'))
            if source not in sources:
                sources[source] = 0
            sources[source] += 1

        parts = []
        for source, count in sorted(sources.items(), key=lambda x: x[1], reverse=True):
            pct = count / len(self.products) * 100 if self.products else 0
            parts.append(f"{source}占{pct:.1f}%")

        return '、'.join(parts) if parts else '数据不足'

    def _get_competition_conclusion(self, kpi: Dict, rating: str) -> str:
        """获取竞争结论"""
        parts = []
        if kpi['hhi'] > 2500:
            parts.append('市场高度集中')
        elif kpi['cr3'] > 50:
            parts.append('头部品牌优势明显')

        if kpi['new_product_percent'] < 20:
            parts.append('新品空间有限')

        if not parts:
            return '竞争环境相对宽松，存在市场机会'
        return '；'.join(parts) + '，需充分评估竞争风险'

    def _prepare_chart_data(self) -> Dict:
        """准备图表数据"""
        # 尝试从 trend_data.json 读取真实趋势数据
        trend_data = self._load_trend_data()

        # 如果有真实数据，使用真实数据；否则使用模拟数据
        if trend_data and 'sales_trend' in trend_data:
            sales_trend = trend_data['sales_trend']
        else:
            # 模拟数据作为降级方案
            sales_trend = {
                'dates': [f'{i}月前' for i in range(25, 0, -1)],
                'sales': [10000 + i * 500 for i in range(25)]
            }

        if trend_data and 'price_trend' in trend_data:
            price_trend = trend_data['price_trend']
        else:
            price_trend = {
                'dates': [f'{i}月前' for i in range(25, 0, -1)],
                'prices': [300 + i * 2 for i in range(25)]
            }

        # 价格分布
        price_dist = self._get_price_distribution()

        # 评分分布
        rating_dist = self._get_rating_distribution()

        # 品牌份额 Top10
        brand_share = self._get_brand_share_data()

        # 卖家来源
        seller_source = self._get_seller_source_data()

        # 品牌评分趋势 (使用 rating_trend 或模拟数据)
        if trend_data and 'rating_trend' in trend_data:
            rating_trend_raw = trend_data['rating_trend']
            brand_rating_trend = {
                'brands': list(brand_share['brands'][:5]),
                'dates': rating_trend_raw.get('dates', [f'{i}月前' for i in range(25, 0, -1)]),
                'data': {brand: rating_trend_raw.get('ratings', [4.2 + i * 0.01 for i in range(25)])
                        for brand in list(brand_share['brands'][:5])}
            }
        else:
            brand_rating_trend = {
                'brands': list(brand_share['brands'][:5]),
                'dates': [f'{i}月前' for i in range(25, 0, -1)],
                'data': {brand: [4.2 + i * 0.01 for i in range(25)] for brand in list(brand_share['brands'][:5])}
            }

        # Top50 产品
        top50_products = self._get_top50_products()

        return {
            'sales_trend': sales_trend,
            'price_trend': price_trend,
            'price_dist': price_dist,
            'rating_dist': rating_dist,
            'brand_share': brand_share,
            'seller_source': seller_source,
            'brand_rating_trend': brand_rating_trend,
            'top50_products': top50_products
        }

    def _load_trend_data(self) -> Optional[Dict]:
        """从 trend_data.json 加载趋势数据"""
        trend_file = self.output_dir / 'trend_data.json'
        if trend_file.exists():
            try:
                with open(trend_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except Exception as e:
                print(f"读取趋势数据失败: {e}")
        return None

    def _get_price_distribution(self) -> List[Dict]:
        """获取价格分布数据"""
        ranges = [
            {'name': '超低价($0-50)', 'min': 0, 'max': 50},
            {'name': '低价($50-150)', 'min': 50, 'max': 150},
            {'name': '中价($150-300)', 'min': 150, 'max': 300},
            {'name': '高价($300-500)', 'min': 300, 'max': 500},
            {'name': '超高价($500+)', 'min': 500, 'max': float('inf')},
        ]

        result = []
        for r in ranges:
            count = sum(1 for p in self.products if r['min'] <= self._safe_float(p.get('价格', p.get('price', 0))) < r['max'])
            result.append({'value': count, 'name': r['name']})

        return result

    def _get_rating_distribution(self) -> Dict:
        """获取评分分布数据"""
        ranges = [
            {'min': 0, 'max': 3.5, 'label': '低分(<3.5)'},
            {'min': 3.5, 'max': 4.0, 'label': '中低分(3.5-4.0)'},
            {'min': 4.0, 'max': 4.3, 'label': '中等(4.0-4.3)'},
            {'min': 4.3, 'max': 4.7, 'label': '中高分(4.3-4.7)'},
            {'min': 4.7, 'max': 5.0, 'label': '高分(4.7+)'},
        ]

        ranges_data = [r['label'] for r in ranges]
        counts = []

        for r in ranges:
            count = sum(1 for p in self.products if r['min'] <= self._safe_float(p.get('评分', p.get('rating', 0))) < r['max'])
            counts.append(count)

        return {'ranges': ranges_data, 'counts': counts}

    def _get_brand_share_data(self) -> Dict:
        """获取品牌份额数据"""
        brands = {}
        for p in self.products:
            brand = p.get('品牌', p.get('brand', 'Unknown'))
            revenue = self._safe_float(p.get('月销额', p.get('monthly_revenue', 0)))
            if revenue == 0:
                price = self._safe_float(p.get('价格', p.get('price', 0)))
                sales = self._safe_int(p.get('月销量', p.get('monthly_sales', 0)))
                revenue = price * sales
            brands[brand] = brands.get(brand, 0) + revenue

        # 排序并取 Top10
        sorted_brands = sorted(brands.items(), key=lambda x: x[1], reverse=True)[:10]

        total_revenue = sum(share for _, share in sorted_brands) if sorted_brands else 1

        return {
            'brands': [brand for brand, _ in sorted_brands],
            'shares': [round(share / total_revenue * 100, 1) for _, share in sorted_brands]
        }

    def _get_seller_source_data(self) -> List[Dict]:
        """获取卖家来源数据"""
        sources = {}
        for p in self.products:
            source = p.get('卖家来源', p.get('seller_source', '其他'))
            sources[source] = sources.get(source, 0) + 1

        return [{'value': count, 'name': source} for source, count in sorted(sources.items(), key=lambda x: x[1], reverse=True)]

    def _get_top50_products(self) -> List[Dict]:
        """获取 Top50 产品数据"""
        result = []
        for i, p in enumerate(self.products[:50], 1):
            price = self._safe_float(p.get('价格', p.get('price', 0)))
            sales = self._safe_int(p.get('月销量', p.get('monthly_sales', 0)))
            total_revenue = sum(self._safe_float(pp.get('价格', pp.get('price', 0))) * self._safe_int(pp.get('月销量', pp.get('monthly_sales', 0))) for pp in self.products)

            market_share = (price * sales / total_revenue * 100) if total_revenue > 0 else 0

            result.append({
                'asin': p.get('ASIN', p.get('asin', '')),
                'title': p.get('标题', p.get('title', ''))[:80],
                'brand': p.get('品牌', p.get('brand', '')),
                'price': round(price, 2),
                'rating': round(self._safe_float(p.get('评分', p.get('rating', 0))), 1),
                'sales': sales,
                'marketShare': round(market_share, 2)
            })

        return result

    def _create_overview_sheet(self, wb, Font):
        """创建概览工作表"""
        ws = wb.create_sheet("概览")

        # 标题
        ws['A1'] = f"{self._get_category_name()} 品类选品分析报告"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:D1')

        # 日期
        ws['A2'] = f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

        row = 4

        # 五维评分
        ws[f'A{row}'] = "五维评分"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1

        max_scores = {'市场规模': 20, '增长潜力': 25, '竞争烈度': 20, '进入壁垒': 20, '利润空间': 15}
        for key, max_score in max_scores.items():
            score = self.scores.get(key, 0)
            ws[f'A{row}'] = key
            ws[f'B{row}'] = score
            ws[f'C{row}'] = max_score
            ws[f'D{row}'] = f"{score/max_score*100:.1f}%"
            row += 1

        row += 1
        ws[f'A{row}'] = "总分"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = self.scores.get('总分', 0)
        ws[f'C{row}'] = 100

        row += 1
        ws[f'A{row}'] = "评级"
        ws[f'B{row}'] = self.scores.get('评级', '')

    def _create_products_sheet(self, wb, Font, PatternFill):
        """创建产品列表工作表"""
        ws = wb.create_sheet("产品列表")

        # 表头
        headers = ['排名', 'ASIN', '品牌', '标题', '价格', '月销量', '评分']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        # 数据
        for row_idx, product in enumerate(self.products, 2):
            ws.cell(row=row_idx, column=1, value=row_idx - 1)
            ws.cell(row=row_idx, column=2, value=product.get('ASIN', ''))
            ws.cell(row=row_idx, column=3, value=product.get('品牌', ''))
            ws.cell(row=row_idx, column=4, value=product.get('标题', '')[:50])
            ws.cell(row=row_idx, column=5, value=product.get('价格', 0))
            ws.cell(row=row_idx, column=6, value=product.get('月销量', 0))
            ws.cell(row=row_idx, column=7, value=product.get('评分', 0))

    def _create_scores_sheet(self, wb, Font):
        """创建评分工作表"""
        ws = wb.create_sheet("评分详情")

        # 表头
        headers = ['维度', '得分', '满分', '占比']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)

        # 数据
        max_scores = {'市场规模': 20, '增长潜力': 25, '竞争烈度': 20, '进入壁垒': 20, '利润空间': 15}
        for row_idx, (key, max_score) in enumerate(max_scores.items(), 2):
            score = self.scores.get(key, 0)
            ws.cell(row=row_idx, column=1, value=key)
            ws.cell(row=row_idx, column=2, value=score)
            ws.cell(row=row_idx, column=3, value=max_score)
            ws.cell(row=row_idx, column=4, value=f"{score/max_score*100:.1f}%")

    def _get_default_markdown_template(self) -> str:
        """获取默认 Markdown 模板"""
        return """# {{CATEGORY_NAME}} 品类选品分析报告

**生成时间**: {{DATE}} {{TIME}}
**数据来源**: Sorftime MCP

---

## 执行摘要

### 综合评级: {{SCORE_评级}}

### 五维评分

| 维度 | 得分 | 满分 |
|------|------|------|
| 市场规模 | {{SCORE_市场规模}} | 20 |
| 增长潜力 | {{SCORE_增长潜力}} | 25 |
| 竞争烈度 | {{SCORE_竞争烈度}} | 20 |
| 进入壁垒 | {{SCORE_进入壁垒}} | 20 |
| 利润空间 | {{SCORE_利润空间}} | 15 |
| **总分** | **{{SCORE_总分}}** | **100** |

---

## 市场数据

| 指标 | 数值 |
|------|------|
{% for key, value in statistics.items() %}| {{ key }} | {{ value }} |
{% endfor %}

---

## Top 产品列表

{{PRODUCTS_TABLE}}

---

## 数据说明

本报告基于 Sorftime MCP 实时数据生成，所有数据文件保存在 `data/` 目录中。

- `statistics.csv` - 统计数据
- `products.csv` - 产品列表
- `scores.csv` - 评分详情
- `raw_data.json` - 原始 JSON 数据

---

*报告自动生成于 {{DATE}}*
"""

    def _get_default_html_template(self) -> str:
        """获取默认 HTML 模板"""
        return """<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{{CATEGORY_NAME}} 品类选品分析</title>
    <style>
        body { font-family: 'Segoe UI', sans-serif; background: #f5f7fa; padding: 20px; }
        .container { max-width: 1200px; margin: 0 auto; background: white; padding: 30px; border-radius: 10px; }
        h1 { color: #333; border-bottom: 2px solid #667eea; padding-bottom: 10px; }
        .score-card { background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 20px; border-radius: 8px; margin: 20px 0; }
        .score-item { display: flex; justify-content: space-between; padding: 5px 0; }
        table { width: 100%; border-collapse: collapse; margin: 20px 0; }
        th, td { padding: 12px; text-align: left; border-bottom: 1px solid #ddd; }
        th { background: #f8f9fa; font-weight: bold; }
    </style>
</head>
<body>
    <div class="container">
        <h1>{{CATEGORY_NAME}} 品类选品分析报告</h1>
        <p>生成时间: {{DATE}}</p>

        <div class="score-card">
            <h2>五维评分</h2>
            <div id="scores"></div>
        </div>

        <h2>产品列表</h2>
        <table id="products-table">
            <thead>
                <tr>
                    <th>排名</th>
                    <th>ASIN</th>
                    <th>品牌</th>
                    <th>价格</th>
                    <th>月销量</th>
                    <th>评分</th>
                </tr>
            </thead>
            <tbody></tbody>
        </table>
    </div>

    <script>
        const scores = {{SCORES_JSON}};
        const products = {{PRODUCTS_JSON}};

        // 渲染评分
        const scoresDiv = document.getElementById('scores');
        const maxScores = {'市场规模': 20, '增长潜力': 25, '竞争烈度': 20, '进入壁垒': 20, '利润空间': 15};
        for (const [key, score] of Object.entries(scores)) {
            if (key in maxScores) {
                const div = document.createElement('div');
                div.className = 'score-item';
                div.innerHTML = `<span>${key}</span><span>${score}/${maxScores[key]}</span>`;
                scoresDiv.appendChild(div);
            }
        }

        // 渲染产品
        const tbody = document.querySelector('#products-table tbody');
        products.slice(0, 20).forEach((p, i) => {
            const tr = document.createElement('tr');
            tr.innerHTML = `
                <td>${i + 1}</td>
                <td>${p.ASIN || ''}</td>
                <td>${p.品牌 || ''}</td>
                <td>$${(p.价格 || 0).toFixed(2)}</td>
                <td>${(p.月销量 || 0).toLocaleString()}</td>
                <td>${(p.评分 || 0).toFixed(1)}★</td>
            `;
            tbody.appendChild(tr);
        });
    </script>
</body>
</html>
"""


def main():
    """命令行入口"""
    if len(sys.argv) < 2:
        print("用法: python generate_reports.py <JSON数据文件> [输出目录]")
        print("\n示例:")
        print("  python generate_reports.py parsed_data.json")
        print("  python generate_reports.py parsed_data.json ./reports")
        sys.exit(1)

    json_file = sys.argv[1]
    output_dir = sys.argv[2] if len(sys.argv) > 2 else None

    # 读取数据
    with open(json_file, 'r', encoding='utf-8') as f:
        data = json.load(f)

    # 生成报告
    generator = CategoryReportGenerator(data, output_dir)
    results = generator.generate_all()

    print("\n" + "=" * 60)
    print("报告生成完成！")
    print("=" * 60)
    print(f"输出目录: {generator.output_dir}")
    print("\n生成的文件:")
    for format_type, path in results.items():
        print(f"  [{format_type.upper()}] {path}")


if __name__ == "__main__":
    main()
